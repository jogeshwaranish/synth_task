"""Sheet ingestion: AG's training workbook, or its per-tab CSV exports.

Parsers are row-oriented (list of dicts), so the file format is isolated to two
thin loaders: stdlib csv and openpyxl. The take-home was distributed as an
xlsx; Basil's local copy is a per-tab CSV export — both must work. The export
already carries converted units (distance_mi, total_elevation_gain_ft,
average_speed_mph): used as-is, no unit math.

UntrustedText fields (name, device_name, wellness notes) originate in the
sheet — DATA, never instructions. They are encrypted at rest by store/db.py
and must be wrapped via synthesize/prompts.wrap_untrusted() before any prompt.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from config import Settings
from ingest import mapping
from schemas import Activity, BikeSplit, RunSplit, Source, Sport, SwimSplit, WellnessDay
from security import crypto
from store import db

Row = dict[str, str | None]

_ACTIVITIES_TAB = "activities_raw"
_WELLNESS_TAB = "health_raw"
_RUN_SPLITS_TAB = "run_splits_raw"
_BIKE_SPLITS_TAB = "bike_splits_raw"
_SWIM_SPLITS_TAB = "swim_splits_raw"

# Split tabs use float-ish indices ("1.0") and the export's own column names.
_SWIM_CONTEXTS = {"pool", "open_water"}
_DISTANCE_UNITS = {"yd", "m"}


def _clean(row: dict) -> Row:
    # Uniform shape for both formats: str values, blanks/None -> None.
    out: Row = {}
    for k, v in row.items():
        if k is None:
            continue  # csv: cells beyond the header row
        s = None if v is None else str(v).strip()
        out[str(k)] = s if s else None
    return out


def _rows_from_csv(path: str | Path) -> list[Row]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [_clean(r) for r in csv.DictReader(f)]


# Identifier columns: floatifying these ("12345" -> "12345.0") would make the
# same row ingest under different ids from csv vs xlsx, duplicating it.
_ID_COLUMNS = {"activity_id"}


def _norm(key: str | None, v: object) -> object:
    # openpyxl returns integral floats as int (3.0 -> 3); keep the "3.0" form
    # so numeric cells always parse with float() downstream — never int().
    if key in _ID_COLUMNS:
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return str(float(v))
    return v


def _rows_from_xlsx(path: str | Path, tab: str) -> list[Row]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        rows_iter = wb[tab].iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        keys = [None if h is None else str(h) for h in header]
        return [
            _clean({k: _norm(k, v) for k, v in zip(keys, raw)})
            for raw in rows_iter
            if any(v is not None for v in raw)
        ]
    finally:
        wb.close()


def _opt_float(v: str | None) -> float | None:
    return None if v is None else float(v)


def _parse_bool(v: str | None) -> bool:
    # csv export says TRUE/FALSE; xlsx bool cells stringify to True/False.
    return v is not None and v.upper() == "TRUE"


def _local_dt(v: str) -> datetime:
    # Sheet wall-clock format, not zero-padded: "2026-05-14 4:05:28". Naive on
    # purpose — local_date derives from this, never from UTC (join rule).
    return datetime.strptime(v, "%Y-%m-%d %H:%M:%S")


def _utc_dt(v: str | None) -> datetime | None:
    return None if not v else datetime.fromisoformat(v.replace("Z", "+00:00"))


def _fallback_activity_id(start_local: datetime, sport_raw: str | None) -> str:
    # Real exports contain rows with no id (watch-app entries). Deterministic
    # fallback keyed on wall-clock start + sport keeps upserts idempotent across
    # re-syncs; a same-second same-sport collision would be a true duplicate.
    return f"sheet-{start_local:%Y%m%dT%H%M%S}-{sport_raw or 'Other'}"


def parse_activity_rows(rows: list[Row], *, athlete_id: str = "ag") -> list[Activity]:
    out: list[Activity] = []
    for row in rows:
        try:
            start_local = _local_dt(row["start_date_local"])
            out.append(Activity(
                activity_id=row.get("activity_id")
                or _fallback_activity_id(start_local, row.get("sport_type")),
                source=Source.SHEET,
                athlete_id=athlete_id,
                start_local=start_local,
                start_utc=_utc_dt(row.get("start_date_utc")),
                local_date=start_local.date(),
                name=row.get("name") or "",
                sport=Sport.normalize(row.get("sport_type") or "Other"),
                is_trainer=_parse_bool(row.get("trainer")),
                moving_time_sec=float(row.get("moving_time_sec") or 0),
                elapsed_time_sec=_opt_float(row.get("elapsed_time_sec")),
                distance_mi=float(row.get("distance_mi") or 0),
                elevation_gain_ft=_opt_float(row.get("total_elevation_gain_ft")),
                avg_speed_mph=_opt_float(row.get("average_speed_mph")),
                avg_hr=_opt_float(row.get("average_heartrate")),
                max_hr=_opt_float(row.get("max_heartrate")),
                avg_watts=_opt_float(row.get("average_watts")),
                weighted_watts=_opt_float(row.get("weighted_average_watts")),
                kilojoules=_opt_float(row.get("kilojoules")),
                avg_cadence=_opt_float(row.get("average_cadence")),
                suffer_score=_opt_float(row.get("suffer_score")),
                calories=_opt_float(row.get("calories")),
                perceived_exertion=_opt_float(row.get("perceived_exertion")),
                device_name=row.get("device_name"),
            ))
        except (KeyError, TypeError, ValueError, ValidationError) as e:
            rid = row.get("activity_id") or "<missing activity_id>"
            # Loud failure with row identity — never skip rows silently.
            raise ValueError(f"bad sheet activity row {rid!r}: {e}") from e
    return out


def parse_wellness_rows(rows: list[Row], *, athlete_id: str = "ag") -> list[WellnessDay]:
    # Column names are a documented ASSUMPTION (CONTRACT.md open items 1-2):
    # AG's wellness tabs are empty as of June 9; verify when rows arrive.
    out: list[WellnessDay] = []
    for row in rows:
        try:
            out.append(WellnessDay(
                # xlsx date cells stringify as "YYYY-MM-DD 00:00:00"; the date part wins.
                local_date=date.fromisoformat(row["local_date"].split(" ")[0]),
                athlete_id=athlete_id,
                in_bed_hours=_opt_float(row.get("in_bed_hours")),
                asleep_hours=_opt_float(row.get("asleep_hours")),
                snoring=_opt_float(row.get("snoring")),
                rhr=_opt_float(row.get("rhr")),
                hrv=_opt_float(row.get("hrv")),
                body_weight_lb=_opt_float(row.get("body_weight_lb")),
                sauna_mins=_opt_float(row.get("sauna_mins")),
                notes=row.get("notes"),
            ))
        except (KeyError, TypeError, ValueError, ValidationError) as e:
            rid = row.get("local_date") or "<missing local_date>"
            raise ValueError(f"bad sheet wellness row {rid!r}: {e}") from e
    return out


def _split_index(v: str | None) -> int:
    # Export stores indices as floats ("1.0"); the contract wants int >= 1.
    return int(float(v))


def parse_run_splits(rows: list[Row]) -> list[RunSplit]:
    out: list[RunSplit] = []
    for row in rows:
        if not row.get("activity_id"):
            continue  # padding/blank row — a real split must reference an activity
        try:
            out.append(RunSplit(
                activity_id=row["activity_id"],
                split_index=_split_index(row["split_index"]),
                distance_mi=float(row.get("distance_mi") or 0),
                moving_time_sec=float(row.get("moving_time_sec") or 0),
                pace_min_per_mi=_opt_float(row.get("pace_min_per_mi")),
                avg_hr=_opt_float(row.get("avg_hr")),
                max_hr=_opt_float(row.get("max_hr")),
                avg_cadence_run=_opt_float(row.get("avg_cadence_run")),
                elevation_gain_ft=_opt_float(row.get("elevation_gain_ft")),
                is_partial=_parse_bool(row.get("is_partial_split")),
            ))
        except (KeyError, TypeError, ValueError, ValidationError) as e:
            rid = f"{row.get('activity_id')}:{row.get('split_index')}"
            raise ValueError(f"bad run split row {rid!r}: {e}") from e
    return out


def parse_bike_splits(rows: list[Row]) -> list[BikeSplit]:
    out: list[BikeSplit] = []
    for row in rows:
        if not row.get("activity_id"):
            continue  # padding/blank row — a real split must reference an activity
        try:
            out.append(BikeSplit(
                activity_id=row["activity_id"],
                split_index=_split_index(row["split_index"]),
                duration_sec=float(row.get("duration_sec") or 0),
                distance_mi=float(row.get("distance_mi") or 0),
                avg_speed_mph=_opt_float(row.get("avg_speed_mph")),
                avg_hr=_opt_float(row.get("avg_hr")),
                avg_power=_opt_float(row.get("avg_power")),
                avg_cadence=_opt_float(row.get("avg_cadence")),
                elevation_gain_ft=_opt_float(row.get("elevation_gain_ft")),
                is_partial=_parse_bool(row.get("is_partial_split")),
            ))
        except (KeyError, TypeError, ValueError, ValidationError) as e:
            rid = f"{row.get('activity_id')}:{row.get('split_index')}"
            raise ValueError(f"bad bike split row {rid!r}: {e}") from e
    return out


def parse_swim_splits(rows: list[Row]) -> list[SwimSplit]:
    out: list[SwimSplit] = []
    for row in rows:
        if not row.get("activity_id"):
            continue  # padding/blank row — a real split must reference an activity
        try:
            ctx = row.get("swim_context")
            unit = row.get("distance_unit")
            out.append(SwimSplit(
                activity_id=row["activity_id"],
                split_index=_split_index(row["split_index"]),
                swim_context=ctx if ctx in _SWIM_CONTEXTS else None,
                distance=float(row.get("distance") or 0),
                distance_unit=unit if unit in _DISTANCE_UNITS else "yd",
                duration_sec=float(row.get("duration_sec") or 0),
                pace_sec_per_100=_opt_float(row.get("pace_sec_per_100")),
                stroke_style=row.get("stroke_style"),  # UntrustedText -> encrypted
                swolf=_opt_float(row.get("swolf")),
                avg_hr=_opt_float(row.get("avg_hr")),
            ))
        except (KeyError, TypeError, ValueError, ValidationError) as e:
            rid = f"{row.get('activity_id')}:{row.get('split_index')}"
            raise ValueError(f"bad swim split row {rid!r}: {e}") from e
    return out


def _load_rows(path: Path, tab: str) -> list[Row]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(path, tab)
    return _rows_from_csv(path)


def _sync_splits(path: Path, conn, key: bytes) -> None:
    # Splits live as extra tabs in the activities workbook (xlsx only; a CSV
    # activities export carries no splits). Each tab is optional.
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return
    wb = load_workbook(path, read_only=True)
    try:
        names = set(wb.sheetnames)
    finally:
        wb.close()
    if _RUN_SPLITS_TAB in names:
        db.upsert_run_splits(conn, parse_run_splits(_rows_from_xlsx(path, _RUN_SPLITS_TAB)))
    if _BIKE_SPLITS_TAB in names:
        db.upsert_bike_splits(conn, parse_bike_splits(_rows_from_xlsx(path, _BIKE_SPLITS_TAB)))
    if _SWIM_SPLITS_TAB in names:
        db.upsert_swim_splits(
            conn, parse_swim_splits(_rows_from_xlsx(path, _SWIM_SPLITS_TAB)), key=key
        )


# security(Anish): wellness ingestion no longer hard-codes a tab/column layout —
# real workbooks vary (this export keeps wellness in `daily_summary`, not the
# empty `health_raw`). ingest/mapping.py infers the column mapping (LLM, cached)
# and parses deterministically; notes stays the encrypted injection surface.
def _tabs_preview(path: Path, *, n_samples: int = 3) -> dict[str, "mapping.TabPreview"]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        out: dict[str, mapping.TabPreview] = {}
        for name in wb.sheetnames:
            rows_iter = wb[name].iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                continue
            keys = [None if h is None else str(h) for h in header]
            headers = [k for k in keys if k is not None]
            samples: list[Row] = []
            for raw in rows_iter:
                if not any(v is not None for v in raw):
                    continue
                samples.append(_clean({k: _norm(k, v) for k, v in zip(keys, raw)}))
                if len(samples) >= n_samples:
                    break
            if not samples:
                continue  # an empty tab (e.g. health_raw here) is not a usable source
            out[name] = mapping.TabPreview(headers=headers, samples=samples)
        return out
    finally:
        wb.close()


def _ingest_wellness_mapped(path: Path, s: Settings, key: bytes) -> list[WellnessDay]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        tabs = _tabs_preview(path)
        return mapping.ingest_wellness(
            tabs, lambda tab: _rows_from_xlsx(path, tab), settings=s, key=key,
            athlete_id=s.strava_athlete_id,
        )
    # CSV has no tabs — treat the whole file as a single candidate sheet.
    rows = _rows_from_csv(path)
    headers = list(rows[0].keys()) if rows else []
    tabs = {"__csv__": mapping.TabPreview(headers=headers, samples=rows[:3])}
    return mapping.ingest_wellness(tabs, lambda _tab: rows, settings=s, key=key,
                                   athlete_id=s.strava_athlete_id)


def sync_sheet(s: Settings, conn) -> int:
    """Ingest the configured sheet export into the store. Returns activity count.

    Activities path is required (caller checks configuration); wellness path is
    optional and an absent/empty wellness source is the documented normal case.
    """
    if s.sheet_activities_path is None:
        raise RuntimeError("Set SHEET_ACTIVITIES_PATH in .env")
    key = crypto.load_or_create_key(s.encryption_key_path)  # encrypt PII at rest
    # Strava + sheet are ONE athlete with two sources; both stamp the same
    # athlete_id (provenance lives on the `source` axis). See DECISIONS.md.
    activities = parse_activity_rows(
        _load_rows(Path(s.sheet_activities_path), _ACTIVITIES_TAB),
        athlete_id=s.strava_athlete_id,
    )
    n = db.upsert_activities(conn, activities, key=key)
    _sync_splits(Path(s.sheet_activities_path), conn, key)  # run/bike/swim split tabs
    if s.sheet_wellness_path is not None and Path(s.sheet_wellness_path).exists():
        days = _ingest_wellness_mapped(Path(s.sheet_wellness_path), s, key)
        db.upsert_wellness(conn, days, key=key)
    return n
